import pandas as pd
import numpy as np
from pathlib import Path
import os
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

import warnings

warnings.filterwarnings(
    "ignore",
    message="Data Validation extension is not supported*"
)

import sys

if getattr(sys, 'frozen', False):
    PATH = Path(sys.executable).parent  # folder chứa exe
else:
    PATH = Path(__file__).parent 


EXPORT_PATH = PATH / 'Export_Data'
TARGET_PATH = PATH / 'HCP_Data'

if not EXPORT_PATH.exists():
    raise FileNotFoundError(f"Folder {EXPORT_PATH} không tồn tại! Vui lòng copy dữ liệu vào cùng folder với exe.")


def run():
    try:
        title = "Auto Checking Target List vs Profiling file"
        width = 50

        print("\n" + "+" + "-"*(width-2) + "+")
        print("|" + title.center(width-2) + "|")
        print("+" + "-"*(width-2) + "+")

        print("Đọc file Profiling_Pivot.xlsx trong folder Export_Data")
        df_pro = pd.read_excel(EXPORT_PATH/'Profiling_Pivot.xlsx')

        print("Đọc File Target_List trong folder TargetList")
        df_target = pd.read_excel(TARGET_PATH / 'Target_List.xlsx', sheet_name='TargetList')

        print("Cleaning and Summary Data")

        df_pro = df_pro.drop_duplicates()
        df_target = df_target.drop_duplicates()

        df_target['Bline'] = np.where(df_target['Source.Name'].str.contains('HA8'), 'ALLIANCE 8',
                             np.where(df_target['Source.Name'].str.contains('HECA 1'), 'ALLIANCE 1',
                             np.where(df_target['Source.Name'].str.contains('HECA 3'), 'ALLIANCE 3',
                             np.where(df_target['Source.Name'].str.contains('HECA6'), 'ALLIANCE 6',
                             np.where(df_target['Source.Name'].str.contains('OB'), 'OWN BRAND', None
                                    )))))
        
        cols = ['Bline','cont_code']
        df_target = df_target[['Bline','cont_code']]

        df_pro = df_pro[['Bline','cont_code']]

        for df in [df_target, df_pro]:
                    for col in df.select_dtypes(include='object'):
                        df[col] = df[col].str.strip()
                    df.replace(['nan', 'None'], None, inplace=True)

        df_target = df_target[(df_target['cont_code'].notna())&(df_target['cont_code'] != 'VN00000000')][['Bline','cont_code']]
        df_pro = df_pro[(df_pro['cont_code'].notna())&(df_pro['cont_code'] != 'VN00000000')][['Bline','cont_code']]

        df_target = df_target.drop_duplicates()
        df_pro = df_pro.drop_duplicates()
        df_target['visit_check'] = 1
        df_pro['visit_check'] = 1

        df_pro_out = df_pro[['Bline','cont_code']].merge(df_target, how = 'left')
        df_target_out = df_target[['Bline','cont_code']].merge(df_pro, how = 'left')

        df_target_count = df_target.groupby('Bline').agg(total_target=('cont_code', 'nunique')).reset_index()
        df_pro_count = df_pro.groupby('Bline').agg(total_profiling=('cont_code', 'nunique')).reset_index()

        df_summ = df_target_count.merge(df_pro_count, how = 'left')

        df_target_inlist = df_target_out.groupby(['Bline'])['visit_check'].sum().reset_index(name = 'total_in_list')



        DF_SUMMARY = df_summ.merge(df_target_inlist, how = 'left')
        DF_SUMMARY['target_vs_profiling'] = (DF_SUMMARY['total_profiling'] / DF_SUMMARY['total_target']).round(2)
        DF_SUMMARY['target_vs_inlist'] = (DF_SUMMARY['total_in_list'] / DF_SUMMARY['total_target']).round(2)
        
        DF_TARGET_LIST = df_target_out.fillna(0)
        DF_PRO_LIST =  df_pro_out[df_pro_out['visit_check'].isna()].fillna(1)
        
        print("Chuẩn hóa xong, tiến hành xuất Excel")

        COLUMN_WIDTHS = {
             1:20, 2: 20, 3: 20, 4: 25, 5: 25, 6:25
        }
        SHEETS_CONFIG = [
             { "sheet_name": "Summary",
                "df": DF_SUMMARY,
                "column_widths": COLUMN_WIDTHS
             },
             {"sheet_name": "Target_List",
              "df": DF_TARGET_LIST,
              "column_widths": COLUMN_WIDTHS
             },
            {"sheet_name": "Outer_Profiling",
              "df": DF_PRO_LIST,
              "column_widths": COLUMN_WIDTHS
             }
                         
        ]

        BLUE_FILL = PatternFill("solid", fgColor="5B9BD5")

        output_file = EXPORT_PATH / "Target_vs_Profiling.xlsx"

        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

            for cfg in SHEETS_CONFIG:
                df = cfg["df"]
                sheet_name = cfg["sheet_name"]
                column_widths = cfg["column_widths"]

                # ----- Write DF -----
                df.to_excel(writer, index=False, sheet_name=sheet_name)
                ws = writer.sheets[sheet_name]

                # ----- Apply header format -----
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.fill = BLUE_FILL
                    if col_idx in column_widths:
                        ws.column_dimensions[cell.column_letter].width = column_widths[col_idx]

                # ----- Sheet setup -----
                ws.row_dimensions[1].height = 30
                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions

        print("✅ Export xong 3 sheet cho data checking Target List vs Profiling")

        print("Tool đã chạy xong. Hãy kiểm tra lại kết quả.")

    except Exception as e:
        print("\n❌ Có lỗi xảy ra:")
        print(e)

        import traceback
        traceback.print_exc()

        input("\nNhấn Enter để thoát...")

if __name__ == "__main__":
        run()