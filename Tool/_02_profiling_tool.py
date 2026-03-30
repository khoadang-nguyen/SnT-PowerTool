import pandas as pd
import numpy as np
from pathlib import Path
import os
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from tqdm import tqdm

import warnings

warnings.filterwarnings(
    "ignore",
    message="Data Validation extension is not supported*"
)


import sys

if getattr(sys, 'frozen', False):
    PATH = Path(sys.executable).parent  # folder chứa exe
else:
    PATH = Path(__file__).parent 


PROFILING_PATH = PATH / 'Profiling'

if not PROFILING_PATH.exists():
    raise FileNotFoundError(f"Folder {PROFILING_PATH} không tồn tại! Vui lòng copy dữ liệu vào cùng folder với exe.")


def run(): 
    try:
        title = "Auto Pivot Profiling file"
        width = 35

        print("\n" + "+" + "-"*(width-2) + "+")
        print("|" + title.center(width-2) + "|")
        print("+" + "-"*(width-2) + "+")


        dfs = []
        files = list(PROFILING_PATH.iterdir())  # chuyển sang list để tqdm biết tổng số
        for file in tqdm(files, desc="Đang xử lý file Profiling", unit="file"):
            tqdm.write(f"Đang đọc: {file.name}")
            if file.is_file() and file.suffix in ['.xlsx', '.xls', '.xlsm', '.xlsb']:
                try:
                    df1 = pd.read_excel(
                        file,
                        sheet_name= 'Market_Survey',
                        engine='openpyxl'
                    )
                    df1 = df1[df1['cont_code'] != 'VN00000000']
                    dfs.append(df1)

                except Exception as e:
                    tqdm.write(f"Lỗi file {file.name}: {e}")
            
            

        df = pd.concat(dfs, ignore_index=True)
        print("Tiến hành clean và pivot data")

        df['sub_ques_name'] = (
        # (np.where(df['sub_ques_name'] == '2. Vui lòng đánh giá từng kênh theo thang điểm từ 0 đến 10, cho hình thức Gặp mặt trực tiếp/ Hội thảo đào tạo y khoa trực tiếp' , '1. Gặp mặt trực tiếp/ Hội nghị đào tạo y khoa trực tiếp.',
        #  np.where(df['sub_ques_name'] == '2. Vui lòng đánh giá từng kênh theo thang điểm từ 0 đến 10, cho hình thức Email/ Mạng xã hội/ Tin nhắn', '2. Email/ Mạng xã hội/ Tin nhắn.',
        #  np.where(df['sub_ques_name'] == '2. Vui lòng đánh giá từng kênh theo thang điểm từ 0 đến 10, cho hình thức Video call/ Hội thảo trực tuyến', '3. Hội thảo trực tuyến/ Gọi Video' ,
        np.where(df['sub_ques_name'] == '1. Trong 6 tháng vừa qua, BS đã được tiếp cận hoặc nhận thông tin từ công ty qua những hình thức/ kênh nào?', df['sub_ques_name'].astype(str) + ' - ' + df['criteria_name'].astype(str), df['sub_ques_name'])
        #  )))
        )
        for col in df.select_dtypes(include='object'):
                df[col] = df[col].str.strip()
        df.replace(['nan', 'None'], None, inplace=True)

        df['txn_timestamp'] = df.groupby([
            'Bline','salesrep_code','salesrep_name',
            'txn_date','cust_code','cust_name','cont_code','cont_name',
            'title_code','title_name'
            ])['txn_timestamp'].transform('max')

        df['expec'] = df.groupby([
            'Bline','salesrep_code','salesrep_name',
            'txn_date','cust_code','cust_name','cont_code','cont_name',
            'title_code','title_name'
        ])['expec'].transform(lambda x: x.dropna().iloc[-1] if x.dropna().any() else pd.NA)

        df = df.sort_values('txn_timestamp').drop_duplicates(df[['Bline','salesrep_code','salesrep_name',
                                                    'txn_date','cust_code','cust_name','cont_code','cont_name',
                                                    'title_code','title_name','sub_ques_name','expec']], keep='last')
        index_col = [col for col in df.columns if col not in ['sub_ques_name', 'sub_ques_code', 'criteria_code','criteria_name','txn_no', 'txn_status', 'visit_id',
                                                    'ques_code','ques_name', 'date_val', 'opt_val', 'gen_comment' , 'cancel_reason',
                                                    'reason_code', 'reason_name']]
        df_piv = df.pivot(index=index_col, columns='sub_ques_name', values='criteria_name').reset_index()

        cols = df_piv.columns[-9:-6]
        df_piv[cols] = df_piv[cols].notna().astype(int)


        RENAME_MAP = {
                        "Bline": "Bline",
                        "txn_no": "txn_no",
                        "txn_status": "txn_status",
                        "salesrep_code": "salesrep_code",
                        "salesrep_name": "salesrep_name",
                        "visit_id": "visit_id",
                        "txn_date": "txn_date",
                        "cust_code": "cust_code",
                        "cust_name": "cust_name",
                        "cont_code": "cont_code",
                        "cont_name": "cont_name",
                        "title_code": "title_code",
                        "title_name": "title_name",
                        "ques_desc": "ques_desc",
                        "expec": "expec",
                        "txn_timestamp": "txn_timestamp",

                        "1. Trong 6 tháng vừa qua, BS đã được tiếp cận hoặc nhận thông tin từ công ty qua những hình thức/ kênh nào? - Email/ Mạng xã hội/ Tin nhắn":
                            "Q1 - Email / MXH / SMS",

                        "1. Trong 6 tháng vừa qua, BS đã được tiếp cận hoặc nhận thông tin từ công ty qua những hình thức/ kênh nào? - Gặp mặt trực tiếp/ Hội nghị đào tạo y khoa trực tiếp":
                            "Q1 - F2F / MEM",

                        "1. Trong 6 tháng vừa qua, BS đã được tiếp cận hoặc nhận thông tin từ công ty qua những hình thức/ kênh nào? - Hội thảo trực tuyến/ Gọi Video":
                            "Q1 - Webinar / Video call",

                        "2. Vui lòng đánh giá từng kênh theo thang điểm từ 0 đến 10, cho hình thức Email/ Mạng xã hội/ Tin nhắn":
                            "Q2 - Rating Email / MXH / SMS",

                        "2. Vui lòng đánh giá từng kênh theo thang điểm từ 0 đến 10, cho hình thức Gặp mặt trực tiếp/ Hội thảo đào tạo y khoa trực tiếp":
                            "Q2 - Rating F2F",

                        "2. Vui lòng đánh giá từng kênh theo thang điểm từ 0 đến 10, cho hình thức Video call/ Hội thảo trực tuyến":
                            "Q2 - Rating Webinar",

                        "3. Ngoài các hình thức tiếp cận như gặp mặt trực tiếp/ tương tác trực tuyến hay thông qua các mạng xã hội…,còn nguồn thông tin nào khác mà BS thấy hữu ích cho việc cập nhật y khoa (ví dụ: hiệp hội, trang web hội chuyên ngành,...)?":
                            "Q3 - Other info sources",

                        "4. BS thường ưu tiên tương tác với công ty theo hình thức nào?":
                            "Q4 - Preferred interaction",

                        "5. Trong các yếu tố sau, yếu tố nào ảnh hưởng nhiều nhất đến quyết định điều trị của bác sĩ?":
                            "Q5 - Main decision factor",
                    }
        df_piv = df_piv.rename(columns=RENAME_MAP)

        COLUMN_WIDTHS = {
                        # Metadata
                        1:10,  2:14,  3:20,  4:20,  5:20,
                        6:14,  7:14,  8:16,  9:22,
                        10:16, 11:22, 12:30, 13:20,


                        # Q1
                        14:25, 15:25, 16:25,

                        # Q2 – channels reached
                        17:30, 18:25, 19:25,

                        # Q3 Q4 Q5 
                        20:25, 21:25, 22:25,
                    }
        

        SHEETS_CONFIG = [
                {
                    "sheet_name": "CHANNEL_SURVEY",
                    "df": df_piv,
                    "column_widths": COLUMN_WIDTHS,

                }]


        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        blue_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")


        print('Tiến hành export data vào folder Profiling')

        output_file = PATH / "Export_Data/Profiling_Pivot.xlsx"

        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            for sheet in SHEETS_CONFIG:
                sheet_name = sheet["sheet_name"]
                df = sheet["df"]
                column_widths = sheet.get("column_widths", {})

                # Write data
                df.to_excel(writer, sheet_name=sheet_name, index=False)

                ws = writer.sheets[sheet_name]

                # 🔥 Freeze header
                ws.freeze_panes = "A2"

                # 🔥 Auto filter
                ws.auto_filter.ref = ws.dimensions

                # 🔥 Header height
                ws.row_dimensions[1].height = 25

                # 🔥 Format header (KHÔNG đụng width)
                for col_idx, col_name in enumerate(df.columns, 1):
                    cell = ws.cell(row=1, column=col_idx)

                    # Header style
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(
                        horizontal='center',
                        vertical='center',
                        wrap_text=True
                    )

                    # Color
                    cell.fill = yellow_fill if col_idx in [5, 7] else blue_fill
                    
                    # Width theo index
                    width = column_widths.get(col_idx, 15)
                    ws.column_dimensions[cell.column_letter].width = width
        print("Tool đã chạy xong. Hãy kiểm tra lại kết quả.")

    except Exception as e:
        print("\n❌ Có lỗi xảy ra:")
        print(e)

        import traceback
        traceback.print_exc()

        input("\nNhấn Enter để thoát...")

if __name__ == "__main__":
    run()