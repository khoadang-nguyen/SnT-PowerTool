import pandas as pd
import numpy as np
from utils.sqlpocket import DBClient
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from tqdm import tqdm
from pathlib import Path
import warnings
warnings.filterwarnings("ignore")

import sys

# Lấy folder chứa file .exe hoặc script
if getattr(sys, 'frozen', False):
    PATH = Path(sys.executable).parent  # folder chứa exe
else:
    PATH = Path(__file__).parent 

db = DBClient()
HCP_QUERY = "select HCPCode, HCPName, Status, Title, Speciality1, Speciality2, HCOCode, HCOName from T_HCP_Master"

def run():
    try:
        title = "Auto auto checking HCP vs Target and Survey file"
        width = 55

        print("\n" + "+" + "-"*(width-2) + "+")
        print("|" + title.center(width-2) + "|")
        print("+" + "-"*(width-2) + "+")

        def get_hcp_data():
            file_path = PATH / 'HCP_Data/HCP_Data.csv'
            
            try:
                # 1. Thử lấy dữ liệu mới từ SQL
                df = db.sql_read_query(HCP_QUERY)
                
                # 2. Nếu thành công, ghi đè vào file CSV để làm "vốn" cho lần sau
                df.to_csv(file_path, index=False)
                print("\nĐã cập nhật dữ liệu mới nhất từ SQL vào HCP_Data.csv")
                return df
                
            except Exception as e:
                # 3. Nếu SQL "ngỏm", đọc ngay file CSV vừa được update từ lần chạy thành công trước đó
                print(f"Không thể kết nối SQL ({e}). Đang dùng dữ liệu dự phòng từ lần chạy gần nhất...")
                
                try:
                    return pd.read_csv(file_path)
                except FileNotFoundError:
                    print("Lỗi: Không tìm thấy cả file CSV dự phòng!")
                    return None # Hoặc trả về DataFrame rỗng: pd.DataFrame()
        
        sources = [
            ("HCP Data",get_hcp_data),
            ("Target List", lambda: pd.read_excel(PATH / 'HCP_Data/Target_List.xlsx', sheet_name='TargetList')),
            ("Profiling Pivot", lambda: pd.read_excel(PATH / 'Export_Data/Profiling_Pivot.xlsx')),
            ("OB BETTEX", lambda: pd.read_excel(PATH / 'Export_Data/OB_Survey_Export.xlsx', sheet_name='BETTEX')),
            ("OB TSC", lambda: pd.read_excel(PATH / 'Export_Data/OB_Survey_Export.xlsx', sheet_name='TSC')),
            ("OB WH", lambda: pd.read_excel(PATH / 'Export_Data/OB_Survey_Export.xlsx', sheet_name='WH'))
        ]

        dfs = {}  # Lưu kết quả với key = name

        for name, loader in tqdm(sources, desc="Đang load dữ liệu", unit="file"):
            try:
                tqdm.write(f"Đang load: {name}")
                df = loader()
                dfs[name] = df
            except Exception as e:
                tqdm.write(f"Lỗi khi load {name}: {e}")
        
        df_hcp = dfs['HCP Data']
        df_target = dfs['Target List']
        df_profiling = dfs['Profiling Pivot']
        df_ob_bettex = dfs['OB BETTEX']
        df_ob_tsc = dfs['OB TSC']
        df_ob_wh = dfs['OB WH']

        print('Đã load xong các data cần thiết.')
        print('Tiến hành cleaning và checking data')

        OB_BT_LIST = df_ob_bettex[(df_ob_bettex['HCP Contact Code'] != 'VN00000000') & (df_ob_bettex['check']==1)][['HCP Contact Code']]
        OB_TSC_LIST = df_ob_tsc[(df_ob_tsc['HCP Contact Code'] != 'VN00000000') & (df_ob_tsc['check']==1)][['HCP Contact Code']]
        OB_WH_LIST = df_ob_wh[(df_ob_wh['HCP Contact Code'] != 'VN00000000') & (df_ob_wh['check']==1)][['HCP Contact Code']]

        PROFI_LIST = df_profiling[(df_profiling['cont_code'] != 'VN00000000')][['cont_code']].drop_duplicates()
        TARGET_LIST = df_target[(df_target['cont_code'] != 'VN00000000')][['cont_code']].drop_duplicates()

        HCP_LIST = (df_hcp[(df_hcp['HCPCode'] != 'VN00000000') & (df_hcp['Status']!='Removed')]
                            )

        for df in [OB_BT_LIST, OB_TSC_LIST, OB_WH_LIST]:
            df.rename(columns = {'HCP Contact Code':'cont_code'}, inplace = True)


        df_target['Bline'] = np.where(df_target['Source.Name'].str.contains('HA8'), 'ALLIANCE 8',
                                    np.where(df_target['Source.Name'].str.contains('HECA 1'), 'ALLIANCE 1',
                                    np.where(df_target['Source.Name'].str.contains('HECA 3'), 'ALLIANCE 3',
                                    np.where(df_target['Source.Name'].str.contains('HECA6'), 'ALLIANCE 6',
                                    np.where(df_target['Source.Name'].str.contains('OB'), 'OWN BRAND', None
                                            )))))


        TARGET_OB = df_target[(df_target['cont_code'] != 'VN00000000') & (df_target['Bline'] == 'OWN BRAND')][['cont_code']].drop_duplicates()
        TARGET_HA1 = df_target[(df_target['cont_code'] != 'VN00000000') & (df_target['Bline'] == 'ALLIANCE 1')][['cont_code']].drop_duplicates()
        TARGET_HA3 = df_target[(df_target['cont_code'] != 'VN00000000') & (df_target['Bline'] == 'ALLIANCE 3')][['cont_code']].drop_duplicates()
        TARGET_HA6 = df_target[(df_target['cont_code'] != 'VN00000000') & (df_target['Bline'] == 'ALLIANCE 6')][['cont_code']].drop_duplicates()
        TARGET_HA8 = df_target[(df_target['cont_code'] != 'VN00000000') & (df_target['Bline'] == 'ALLIANCE 8')][['cont_code']].drop_duplicates()

        OB_BT_LIST['OB_BETTEX'] = 1
        OB_TSC_LIST['OB_TSC'] = 1
        OB_WH_LIST['OB_WH'] = 1
        PROFI_LIST['PROFILING'] = 1
        TARGET_LIST['TARGET'] = 1
        TARGET_OB['TARGET_OB'] = 1
        TARGET_HA1['TARGET_HA1'] = 1
        TARGET_HA3['TARGET_HA3'] = 1
        TARGET_HA6['TARGET_HA6'] = 1
        TARGET_HA8['TARGET_HA8'] = 1

        OUTER_LIST = TARGET_LIST[~TARGET_LIST['cont_code'].isin(HCP_LIST['HCPCode'].unique())][['cont_code']]

        new_rows = pd.DataFrame({
            "HCPCode": OUTER_LIST["cont_code"]
        })


        for col in HCP_LIST.columns:
            if col != "HCPCode":
                new_rows[col] = pd.NA


        new_rows = new_rows[HCP_LIST.columns]

        HCP_LIST = pd.concat([HCP_LIST, new_rows], ignore_index=True)
        HCP_LIST.rename(columns = {'HCPCode':'cont_code'}, inplace = True)

        DFF = HCP_LIST.merge(TARGET_LIST, how = 'left', on = 'cont_code')
        DFF = DFF.merge(TARGET_OB, how = 'left', on = 'cont_code')
        DFF = DFF.merge(TARGET_HA1, how = 'left', on = 'cont_code')
        DFF = DFF.merge(TARGET_HA3, how = 'left', on = 'cont_code')
        DFF = DFF.merge(TARGET_HA6, how = 'left', on = 'cont_code')
        DFF = DFF.merge(TARGET_HA8, how = 'left', on = 'cont_code')
        DFF = DFF.merge(PROFI_LIST,  how = 'left', on = 'cont_code')
        DFF = DFF.merge(OB_BT_LIST,  how = 'left', on = 'cont_code')
        DFF = DFF.merge(OB_TSC_LIST,  how = 'left', on = 'cont_code')
        DFF = DFF.merge(OB_WH_LIST,  how = 'left', on = 'cont_code')

        DFF.iloc[:,[-1,-2,-3,-4,-5,-6,-7,-8,-9]] = DFF.iloc[:,[-1,-2,-3,-4,-5,-6,-7,-8,-9]].fillna(0)

        DFF['Survey_Check'] = np.where((DFF['PROFILING'] == 1) & (DFF['OB_BETTEX'] + DFF['OB_TSC'] + DFF['OB_WH']) >=1, 1, 0)


        print('Đã hoàn thiện xong bộ data, tiến hành export dữ liệu vào Folder Export')


        # ---------- COLUMN WIDTHS ----------

        COLUMN_WIDTHS = {1: 20, 2: 25, 4: 25, 5: 30, 6: 30, 8: 30} 
        DEFAULT_WIDTH = 20

        GRAY_COLUMNS = { i for i in range(8,19)}

        # ---------- STYLE ----------
        BLUE_FILL = PatternFill("solid", fgColor="5B9BD5")
        GRAY_FILL = PatternFill("solid", fgColor="D9D9D9")

        FONT_WHITE = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        FONT_BLACK = Font(name="Calibri", bold=True, size=10, color="000000")

        ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
        THIN = Side(style="thin")
        BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

        # -------------------------------------
        # EXPORT EXCEL
        # -------------------------------------

        output_file = PATH / "Export_Data/HCP_Vs_Everything.xlsx"

        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            # Ghi DataFrame
            DFF.to_excel(writer, index=False, sheet_name="Survey")
            ws = writer.sheets["Survey"]

            # ---------- HEADER STYLE + COLUMN WIDTH ----------
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_idx)

                # style header
                is_gray = col_idx in GRAY_COLUMNS
                cell.fill = GRAY_FILL if is_gray else BLUE_FILL
                cell.font = FONT_BLACK if is_gray else FONT_WHITE
                cell.alignment = ALIGNMENT
                cell.border = BORDER

                # set column width: dùng COLUMN_WIDTHS nếu có, còn không dùng DEFAULT_WIDTH
                ws.column_dimensions[cell.column_letter].width = COLUMN_WIDTHS.get(col_idx, DEFAULT_WIDTH)

            # ---------- SHEET SETUP ----------
            ws.row_dimensions[1].height = 70  # header cao
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

        print("✅ Export xong data checking list survey, target so với HCP list")
        print("Tool đã chạy xong. Hãy kiểm tra lại kết quả.")

    except Exception as e:
        print("\n❌ Có lỗi xảy ra:")
        print(e)

        import traceback
        traceback.print_exc()

        input("\nNhấn Enter để thoát...")

if __name__ == "__main__":
    run()