import warnings
import pandas as pd
from pathlib import Path
import numpy as np
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from tqdm import tqdm

import warnings

warnings.filterwarnings(
    "ignore",
    message="Data Validation extension is not supported*"
)


import sys

# Lấy folder chứa file .exe hoặc script
if getattr(sys, 'frozen', False):
    PATH = Path(sys.executable).parent  # folder chứa exe
else:
    PATH = Path(__file__).parent        # chạy script bình thường

LEO_PATH = PATH / 'HA8_LEO'
OB_PATH = PATH / 'OB'

# Kiểm tra folder có tồn tại không
if not LEO_PATH.exists():
    raise FileNotFoundError(f"Folder {LEO_PATH} không tồn tại! Vui lòng copy dữ liệu vào cùng folder với exe.")

if not OB_PATH.exists():
    raise FileNotFoundError(f"Folder {OB_PATH} không tồn tại! Vui lòng copy dữ liệu vào cùng folder với exe.")

def run():
    try:
        title = "Auto concat file"
        width = 30

        print("\n" + "+" + "-"*(width-2) + "+")
        print("|" + title.center(width-2) + "|")
        print("+" + "-"*(width-2) + "+")

        title = "LEO - Process"
        width = 40

        print("\n" + "+" + "-"*(width-2) + "+")
        print("|" + title.center(width-2) + "|")
        print("+" + "-"*(width-2) + "+")


        dfs = []
        files = list(LEO_PATH.iterdir())  # chuyển sang list để tqdm biết tổng số
        for file in tqdm(files, desc="Đang xử lý file LEO", unit="file"):
            tqdm.write(f"Đang đọc: {file.name}")
            if file.is_file() and file.suffix in ['.xlsx', '.xls', '.xlsm', '.xlsb']:
                try:
                    df = pd.read_excel(
                        file,
                        sheet_name="LEO",
                        header=11,
                        usecols='A:AW',
                        engine='openpyxl'
                    )
                    df = df[df['HCP Contact Code'] != 'VN00000000']
                    dfs.append(df)

                except Exception as e:
                    tqdm.write(f"Lỗi file {file.name}: {e}")
            
            

        LEO_DF = pd.concat(dfs, ignore_index=True)
        print("Đã nối xong toàn bộ file")
        print("Tiến hành clean và tạo cột score")

        for col in LEO_DF.select_dtypes(include='object'):
            LEO_DF[col] = LEO_DF[col].str.strip()
        LEO_DF.replace(['nan', 'None','null'], None, inplace=True)


        LEO_DF = LEO_DF[~LEO_DF['Rep Code'].isna()].drop_duplicates()
        #rating and check valid
        LEO_DF['Potential'] = np.where(LEO_DF.iloc[:, 26].isna(),0,1)
        LEO_DF['Brand1'] = np.where((LEO_DF.iloc[:, 28].notna()) | (LEO_DF.iloc[:, 29].notna()) | (LEO_DF.iloc[:, 30].notna()) | (LEO_DF.iloc[:, 31].notna()), 1, 0)
        LEO_DF['Brand2'] = np.where((LEO_DF.iloc[:, 33].notna()) | (LEO_DF.iloc[:, 34].notna()) | (LEO_DF.iloc[:, 35].notna()) | (LEO_DF.iloc[:, 36].notna()), 1, 0)
        LEO_DF['Brand3'] = np.where((LEO_DF.iloc[:, 38].notna()) | (LEO_DF.iloc[:, 39].notna()) | (LEO_DF.iloc[:, 40].notna()) | (LEO_DF.iloc[:, 41].notna()), 1, 0)

        LEO_DF['check'] = np.where(LEO_DF['Potential'] == 0, 0,
                np.where(LEO_DF['Potential'] == 1 & ((LEO_DF['Brand1'] + LEO_DF['Brand2'] + LEO_DF['Brand3']) == 3), 1,0))
        
        LEO_DF['valid_check'] = np.where ((LEO_DF.iloc[:, 26] * 3 >= (LEO_DF.iloc[:, 29].fillna(0).astype(int) + LEO_DF.iloc[:, 34].fillna(0).astype(int) + LEO_DF.iloc[:, 39].fillna(0).astype(int)))
                                & (LEO_DF['check'] == 1),1,0
                                )

        COLUMN_WIDTHS = {
            # Master data
            1: 15, 2: 10, 3: 12, 4: 10,
            5: 22, 6: 18, 7: 18, 8: 22,
            9: 22, 10: 15, 11: 20, 12: 22,
            13: 20, 14: 20, 15: 22,
            16: 15, 17: 22,

            # Channel question block
            18: 45,
            19: 8, 20: 10, 21: 10, 22: 14, 23: 18,

            24: 45,
            25: 35,
            26: 35,
            27: 30,

            # Brand 1
            28: 12,
            29: 40, 30: 35, 31: 35, 32: 40,

            # Brand 2
            33: 12,
            34: 40, 35: 35, 36: 38, 37: 40,

            # Brand 3
            38: 12,
            39: 40, 40: 40, 41: 40, 42: 40,

            # Status
            43: 14, 44: 14, 45: 14, 46: 14,
            47: 10, 48: 10, 49: 10,

            # Score
            50: 10
        }
        GRAY_COLUMNS = {13, 14, 28, 33, 38}


        output_file = PATH / "Export_Data/LEO_Survey_Export.xlsx"
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            LEO_DF.to_excel(writer, index=False, sheet_name="Survey")
            ws = writer.sheets["Survey"]

            # ---------- STYLE ----------
            BLUE_FILL = PatternFill("solid", fgColor="5B9BD5")
            GRAY_FILL = PatternFill("solid", fgColor="D9D9D9")

            FONT_WHITE = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
            FONT_BLACK = Font(name="Calibri", bold=True, size=10, color="000000")

            ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
            THIN = Side(style="thin")
            BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

            # ---------- APPLY HEADER ----------
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_idx)

                is_gray = col_idx in GRAY_COLUMNS
                cell.fill = GRAY_FILL if is_gray else BLUE_FILL
                cell.font = FONT_BLACK if is_gray else FONT_WHITE
                cell.alignment = ALIGNMENT
                cell.border = BORDER

                if col_idx in COLUMN_WIDTHS:
                    ws.column_dimensions[cell.column_letter].width = COLUMN_WIDTHS[col_idx]

            # ---------- SHEET SETUP ----------
            ws.row_dimensions[1].height = 70
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions


        print("✅ Export xong sheet cho data LEO")

        #==================
        #  OB
        #==================

        title = "OB - Process"
        width = 40

        print("\n" + "+" + "-"*(width-2) + "+")
        print("|" + title.center(width-2) + "|")
        print("+" + "-"*(width-2) + "+")

        dfs_bt = []
        dfs_tsc = []
        dfs_wh = []

        files2 = list(OB_PATH.iterdir())

        for file in tqdm(files2, desc="Đang xử lý OB", unit="file"):
            tqdm.write(f"Đang đọc: {file.name}")
            if file.is_file() and file.suffix in ['.xlsx', '.xls', '.xlsm', '.xlsb']:
                try:
                    df1 = pd.read_excel(file, sheet_name="BETEX", header=11, usecols='A:AD')
                    df1 = df1[df1['HCP Contact Code'] != 'VN00000000']

                    df2 = pd.read_excel(file, sheet_name="TSC", header=11, usecols='A:AG')
                    df2 = df2[df2['HCP Contact Code'] != 'VN00000000']

                    df3 = pd.read_excel(file, sheet_name="WH", header=11, usecols='A:AL')
                    df3 = df3[df3['HCP Contact Code'] != 'VN00000000']

                    dfs_bt.append(df1)
                    dfs_tsc.append(df2)
                    dfs_wh.append(df3)

                except Exception as e:
                    tqdm.write(f"Lỗi file {file.name}: {e}")

        OB_BETTEX_DF = pd.concat(dfs_bt, ignore_index=True)
        OB_TSC_DF = pd.concat(dfs_tsc, ignore_index=True)
        OB_WH_DF = pd.concat(dfs_wh, ignore_index=True)

        print("Đã nối xong toàn bộ file")
        print("Tiến hành clean và tạo cột score")

        OB_BETTEX_DF = OB_BETTEX_DF[OB_BETTEX_DF['Rep Name'].notna()].drop_duplicates()
        OB_TSC_DF = OB_TSC_DF[OB_TSC_DF['Rep Name'].notna()].drop_duplicates()
        OB_WH_DF = OB_WH_DF[OB_WH_DF['Rep Name'].notna()].drop_duplicates()

        for df in [OB_BETTEX_DF, OB_TSC_DF, OB_WH_DF]:
            for col in df.select_dtypes(include='object'):
                df[col] = df[col].str.strip()
            df.replace(['nan', 'None'], None, inplace=True)

        
        #Valid and check
        OB_BETTEX_DF['Potential'] = np.where(OB_BETTEX_DF.iloc[:, 19].isna(),0,1)
        OB_BETTEX_DF['Brand1'] = np.where((OB_BETTEX_DF.iloc[:, 21].notna()) | (OB_BETTEX_DF.iloc[:, 22].notna()) | (OB_BETTEX_DF.iloc[:, 23].notna()), 1, 0)
        OB_BETTEX_DF['Brand2'] = np.where((OB_BETTEX_DF.iloc[:, 25].notna()) | (OB_BETTEX_DF.iloc[:, 26].notna()) | (OB_BETTEX_DF.iloc[:, 27].notna()), 1, 0)


        OB_BETTEX_DF['check'] = np.where(OB_BETTEX_DF['Potential'] == 0, 0,
                        np.where(OB_BETTEX_DF['Potential'] == 1 & ((OB_BETTEX_DF['Brand1'] + OB_BETTEX_DF['Brand2']) == 2), 1,0))

        OB_BETTEX_DF['valid_check'] = np.where ((OB_BETTEX_DF.iloc[:, 19] * 2 >= (OB_BETTEX_DF.iloc[:, 22].fillna(0).astype(int) + OB_BETTEX_DF.iloc[:, 26].fillna(0).astype(int)))
                                        & (OB_BETTEX_DF['check'] == 1),1,0
        )


        OB_WH_DF['Potential'] = np.where(OB_WH_DF.iloc[:, 19].isna(),0,1)
        OB_WH_DF['Brand1'] = np.where((OB_WH_DF.iloc[:, 21].notna()) | (OB_WH_DF.iloc[:, 22].notna()) | (OB_WH_DF.iloc[:, 23].notna()), 1, 0)
        OB_WH_DF['Brand2'] = np.where((OB_WH_DF.iloc[:, 25].notna()) | (OB_WH_DF.iloc[:, 26].notna()) | (OB_WH_DF.iloc[:, 27].notna()) | (OB_WH_DF.iloc[:, 28].notna()) | (OB_WH_DF.iloc[:, 29].notna()), 1, 0)
        OB_WH_DF['Brand3'] = np.where((OB_WH_DF.iloc[:, 31].notna()) | (OB_WH_DF.iloc[:, 32].notna()) | (OB_WH_DF.iloc[:, 33].notna()) | (OB_WH_DF.iloc[:, 34].notna()), 1, 0)


        OB_WH_DF['check'] = np.where(OB_WH_DF['Potential'] == 0, 0,
                        np.where(OB_WH_DF['Potential'] == 1 & ((OB_WH_DF['Brand1'] + OB_WH_DF['Brand2']) == 2), 1,0))

        OB_WH_DF['valid_check'] = np.where ((OB_WH_DF.iloc[:, 19] * 3 >= (OB_WH_DF.iloc[:, 22].fillna(0).astype(int) + OB_WH_DF.iloc[:, 26].fillna(0).astype(int)) + OB_WH_DF.iloc[:, 32].fillna(0).astype(int))
                                        & (OB_WH_DF['check'] == 1),1,0
        )

        OB_TSC_DF['Potential'] = np.where(OB_TSC_DF.iloc[:, 20].isna(),0,1)
        OB_TSC_DF['Brand1'] = np.where((OB_TSC_DF.iloc[:, 22].notna()) | (OB_TSC_DF.iloc[:, 23].notna()) | (OB_TSC_DF.iloc[:, 24].notna()) | (OB_TSC_DF.iloc[:, 25].notna()), 1, 0)
        OB_TSC_DF['Brand2'] = np.where((OB_TSC_DF.iloc[:, 27].notna()) | (OB_TSC_DF.iloc[:, 28].notna()) | (OB_TSC_DF.iloc[:, 29].notna()) | (OB_TSC_DF.iloc[:, 30].notna()), 1, 0)


        OB_TSC_DF['check'] = np.where(OB_TSC_DF['Potential'] == 0, 0,
                        np.where(OB_TSC_DF['Potential'] == 1 & ((OB_TSC_DF['Brand1'] + OB_TSC_DF['Brand2']) == 2), 1,0))

        OB_TSC_DF['valid_check'] = np.where ((OB_TSC_DF.iloc[:, 20] * 2 >= (OB_TSC_DF.iloc[:, 23].fillna(0).astype(int) + OB_TSC_DF.iloc[:, 28].fillna(0).astype(int)))
                                        & (OB_TSC_DF['check'] == 1),1,0
        )


        COLUMN_WIDTHS_BT = {
            # Master data
            1: 15, 2: 10, 3: 12, 4: 10,
            5: 20, 6: 15, 7: 16, 8: 14,
            9: 18, 10: 18, 11: 22,
            12: 20, 13: 20, 14: 22,
            15: 20, 16: 20, 17: 22,
            18: 15, 19: 22,

            # Question block
            20: 45,

            # Brand 1
            21: 12,
            22: 40, 23: 40, 24: 40,

            # Brand 2
            25: 12,
            26: 40, 27: 40, 28: 40,

            # Status
            29: 14,
            30: 14,
            31: 10
        }
        GRAY_BT = {20, 24}

        COLUMN_WIDTHS_TSC = {
            # Master data
            1:15,  2:10,  3:12,  4:10,
            5:20,  6:15,  7:16,  8:14,
            9:12, 10:18, 11:18, 12:22,
            13:20, 14:20, 15:22, 16:20,
            17:20, 18:22, 19:15, 20:22,

            # Question block
            21:45,

            # Brand 1
            22:12,
            23:40, 24:40, 25:40, 26:40,

            # Brand 2
            27:12,
            28:40, 29:40, 30:40, 31:40,

            # Status
            32:14,
            33:14,
            34:10
        }
        GRAY_TSC = {21, 26}

        COLUMN_WIDTHS_WH = {
            # Master data
            1:15,  2:10,  3:12,  4:10,
            5:20,  6:15,  7:16,  8:14,
            9:18, 10:18, 11:22,
            12:20, 13:20, 14:22,
            15:20, 16:20, 17:22,
            18:15, 19:22,

            # Question opening
            20:45,

            # Brand 1
            21:12,
            22:40, 23:40, 24:40,

            # Brand 2
            25:12,
            26:40, 27:35, 28:40, 29:40, 30:40,

            # Brand 3
            31:12,
            32:40, 33:40, 34:40, 35:40,

            # Status
            36:14, 37:14, 38:14,
            39: 10
        }

        GRAY_WH = {20,24,30}

        output_file = PATH / "Export_Data/OB_Survey_Export.xlsx"

        # ==================================================
        # 1. KHAI BÁO STYLE (DÙNG CHUNG)
        # ==================================================
        BLUE_FILL = PatternFill("solid", fgColor="5B9BD5")
        GRAY_FILL = PatternFill("solid", fgColor="D9D9D9")

        FONT_WHITE = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        FONT_BLACK = Font(name="Calibri", bold=True, size=10, color="000000")

        ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
        THIN = Side(style="thin")
        BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

        # ==================================================
        # 2. CONFIG CHO TỪNG SHEET (THEO SỐ)
        # ==================================================
        SHEETS_CONFIG = [
            {
                "sheet_name": "BETTEX",
                "df": OB_BETTEX_DF,
                "column_widths": COLUMN_WIDTHS_BT,
                "gray_columns": GRAY_BT,
            },
            {
                "sheet_name": "TSC",
                "df": OB_TSC_DF,
                "column_widths": COLUMN_WIDTHS_TSC,
                "gray_columns": GRAY_TSC,
            },
            {
                "sheet_name": "WH",
                "df": OB_WH_DF,
                "column_widths": COLUMN_WIDTHS_WH,
                "gray_columns": GRAY_WH,
            },
        ]

        # ==================================================
        # 3. EXPORT + APPLY FORMAT
        # ==================================================
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

            for cfg in SHEETS_CONFIG:
                df = cfg["df"]
                sheet_name = cfg["sheet_name"]
                column_widths = cfg["column_widths"]
                gray_columns = cfg["gray_columns"]

                # ----- Write DF -----
                df.to_excel(writer, index=False, sheet_name=sheet_name)
                ws = writer.sheets[sheet_name]

                # ----- Apply header format -----
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=1, column=col_idx)

                    is_gray = col_idx in gray_columns
                    cell.fill = GRAY_FILL if is_gray else BLUE_FILL
                    cell.font = FONT_BLACK if is_gray else FONT_WHITE
                    cell.alignment = ALIGNMENT
                    cell.border = BORDER

                    if col_idx in column_widths:
                        ws.column_dimensions[cell.column_letter].width = column_widths[col_idx]

                # ----- Sheet setup -----
                ws.row_dimensions[1].height = 70
                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions

        print("✅ Export xong 3 sheet cho data OB")

        print("Tool đã chạy xong. Hãy kiểm tra lại kết quả.")

    except Exception as e:
        print("\n❌ Có lỗi xảy ra:")
        print(e)

        import traceback
        traceback.print_exc()

        input("\nNhấn Enter để thoát...")

if __name__ == "__main__":
    run()